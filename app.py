import os, csv, re, time, uuid, threading
import requests
from flask import Flask, render_template, request, jsonify, send_file
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
os.makedirs('uploads', exist_ok=True)
os.makedirs('outputs', exist_ok=True)

SAVE_EVERY = 20
DELAY = 0.3
jobs = {}

HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36','Accept':'text/html,application/xhtml+xml,*/*;q=0.8','Accept-Language':'en-US,en;q=0.5'}
EMAIL_REGEX = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
SKIP_DOMAINS = {'wixpress.com','sentry.io','example.com','domain.com','yourdomain.com','wordpress.com','google.com','schema.org','placeholder.com','test.com','amazonaws.com','cloudflare.com','jquery.com','w3.org','gravatar.com','facebook.com','twitter.com','instagram.com','linkedin.com','apple.com','microsoft.com'}
CONTACT_HINTS = ['contact','about','reach','connect','get-in-touch','getintouch','support','help','team','enquiry','inquiry','info']
SKIP_EXT = {'.png','.jpg','.jpeg','.gif','.svg','.webp','.css','.js','.pdf','.mp4'}

def is_valid_email(e):
    e = e.lower().strip()
    if len(e)>80 or '@' not in e: return False
    l,d = e.rsplit('@',1)
    if len(l)<1 or len(d)<4: return False
    for s in SKIP_DOMAINS:
        if d==s or d.endswith('.'+s): return False
    for x in SKIP_EXT:
        if e.endswith(x): return False
    return '/' not in l and '\\' not in l

def get_page(url,timeout=8):
    try:
        import urllib3; urllib3.disable_warnings()
        r=requests.get(url,headers=HEADERS,timeout=timeout,allow_redirects=True,verify=False)
        if r.status_code==200: r.encoding=r.apparent_encoding or 'utf-8'; return r.text
    except: pass
    return None

def extract_emails(html):
    found=set()
    for e in EMAIL_REGEX.findall(html):
        if is_valid_email(e): found.add(e.lower())
    try:
        soup=BeautifulSoup(html,'html.parser')
        for a in soup.find_all('a',href=True):
            h=a['href']
            if h.lower().startswith('mailto:'):
                em=h[7:].split('?')[0].strip()
                if is_valid_email(em): found.add(em.lower())
    except: pass
    return list(found)

def find_contact_links(html,base_url):
    links=[]
    try:
        soup=BeautifulSoup(html,'html.parser'); bd=urlparse(base_url).netloc
        for a in soup.find_all('a',href=True):
            href=a['href'].strip(); text=a.get_text(strip=True).lower(); hl=href.lower()
            if any(h in hl or h in text for h in CONTACT_HINTS):
                full=urljoin(base_url,href); p=urlparse(full)
                if p.netloc==bd and not any(p.path.lower().endswith(x) for x in SKIP_EXT): links.append(full)
    except: pass
    seen=set(); uniq=[]
    for l in links:
        if l not in seen: seen.add(l); uniq.append(l)
    return uniq[:3]

def scrape_website(website):
    if not website or not website.strip(): return []
    url=website.strip()
    if not url.startswith('http'): url='https://'+url
    html=get_page(url) or get_page(url.replace('https://','http://'))
    if not html: return []
    emails=extract_emails(html)
    if not emails:
        for link in find_contact_links(html,url)[:2]:
            time.sleep(0.2); chtml=get_page(link)
            if chtml: emails=extract_emails(chtml)
            if emails: break
    return list(dict.fromkeys(e for e in emails if is_valid_email(e)))[:5]

def save_progress(job_id,results):
    if results:
        with open(os.path.join('outputs',f'{job_id}_progress.csv'),'w',newline='',encoding='utf-8-sig') as f:
            w=csv.DictWriter(f,fieldnames=results[0].keys()); w.writeheader(); w.writerows(results)
    with open(os.path.join('outputs',f'{job_id}_lastrow.txt'),'w') as f: f.write(str(len(results)))

def save_final(job_id,results):
    out_csv=os.path.join('outputs',f'{job_id}.csv'); out_xlsx=os.path.join('outputs',f'{job_id}.xlsx')
    with open(out_csv,'w',newline='',encoding='utf-8-sig') as f:
        if results: w=csv.DictWriter(f,fieldnames=results[0].keys()); w.writeheader(); w.writerows(results)
    try:
        import openpyxl; from openpyxl.styles import PatternFill,Font,Alignment
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Email Results"
        hdrs=list(results[0].keys()) if results else []
        hfill=PatternFill("solid",fgColor="111827"); hfont=Font(color="34D399",bold=True)
        for ci,h in enumerate(hdrs,1):
            c=ws.cell(row=1,column=ci,value=h.upper()); c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal='center')
        efill=PatternFill("solid",fgColor="ECFDF5")
        for ri,row in enumerate(results,2):
            for ci,key in enumerate(hdrs,1):
                cell=ws.cell(row=ri,column=ci,value=row[key])
                if key=='emails' and row[key]: cell.fill=efill; cell.font=Font(color="065F46")
        ws.column_dimensions['A'].width=40; ws.column_dimensions['C'].width=35; ws.column_dimensions['D'].width=50; ws.freeze_panes='A2'
        wb.save(out_xlsx)
    except: pass
    return out_csv,out_xlsx

def run_scraping_job(job_id,rows,delay,start_from=0,existing_results=None):
    job=jobs[job_id]
    job.update({'status':'running','total':len(rows),'results':list(existing_results or []),
                'found':sum(1 for r in (existing_results or []) if r.get('emails','')),'processed':start_from,'log':[],'start_time':time.time()})
    def log(msg,level='info'):
        job['log'].append({'msg':msg,'level':level,'time':time.strftime('%H:%M:%S')})
        if len(job['log'])>200: job['log']=job['log'][-200:]
    if start_from>0: log(f"⏩ Resuming from row {start_from} — {job['found']} emails already found",'success')
    else: log(f"🚀 Started — {len(rows)} rows, delay={delay}s")
    for i,row in enumerate(rows):
        if i<start_from: continue
        if job.get('cancelled'):
            log('⛔ Cancelled — data save ho gaya!','warn')
            job['status']='cancelled'
            save_progress(job_id,job['results'])
            out_csv,out_xlsx=save_final(job_id,job['results'])
            job['out_csv']=out_csv; job['out_xlsx']=out_xlsx
            return
        website=row.get('website','').strip(); title=row.get('title','N/A')
        if website:
            try:
                emails=scrape_website(website); email_str=' | '.join(emails)
                if emails: job['found']+=1; log(f"✓ {title[:40]} → {email_str[:60]}",'success')
                else: log(f"✗ {title[:40]} — no email",'info')
            except Exception as ex: emails=[]; email_str=''; log(f"⚠ {title[:35]} — {str(ex)[:35]}",'warn')
            time.sleep(delay)
        else: emails=[]; email_str=''
        result={'title':title,'phone':row.get('phone',''),'website':website,'emails':email_str,'email_count':len(emails),'city':row.get('city','') or row.get('state',''),'countryCode':row.get('countryCode',''),'totalScore':row.get('totalScore',''),'categoryName':row.get('categoryName','')}
        job['results'].append(result)
        job['processed']=i+1; job['progress']=round((i+1)/len(rows)*100,1)
        if (i+1)%SAVE_EVERY==0:
            save_progress(job_id,job['results'])
            log(f"💾 Auto-saved! {job['processed']}/{len(rows)} — {job['found']} emails",'info')
    out_csv,out_xlsx=save_final(job_id,job['results'])
    for fn in [f'outputs/{job_id}_progress.csv',f'outputs/{job_id}_lastrow.txt']:
        try: os.remove(fn)
        except: pass
    elapsed=int(time.time()-job['start_time'])
    job.update({'status':'done','out_csv':out_csv,'out_xlsx':out_xlsx,'elapsed':elapsed})
    log(f"✅ Done! {job['found']}/{len(rows)} emails in {elapsed//60}m {elapsed%60}s",'success')

# ── STARTUP: Check for any previously saved jobs ──
def restore_saved_jobs():
    if not os.path.exists('outputs'): return
    for fn in os.listdir('outputs'):
        if not fn.endswith('_lastrow.txt'): continue
        job_id=fn.replace('_lastrow.txt','')
        pcv=os.path.join('outputs',f'{job_id}_progress.csv')
        lrf=os.path.join('outputs',fn)
        if not os.path.exists(pcv): continue
        try:
            with open(lrf) as f: last_row=int(f.read().strip() or 0)
            with open(pcv,encoding='utf-8-sig') as f: existing=list(csv.DictReader(f))
            found=sum(1 for r in existing if r.get('emails',''))
            jobs[job_id]={'status':'paused','progress':round(last_row/max(last_row,1)*100,1),'processed':last_row,'total':last_row,'found':found,'results':existing,'log':[{'msg':f'⏸ Incomplete session found — {last_row} rows processed, {found} emails found. Click Resume to continue.','level':'warn','time':time.strftime('%H:%M:%S')}],'cancelled':False,'start_time':time.time(),'out_csv':pcv}
            print(f"  ⏸ Restored paused job: {job_id} ({last_row} rows, {found} emails found)")
        except Exception as e: print(f"  ⚠ Restore failed {job_id}: {e}")

@app.route('/')
def index(): return render_template('index.html')

@app.route('/saved_jobs')
def saved_jobs():
    """Return any paused/saved jobs to the frontend"""
    paused=[]
    for jid,job in jobs.items():
        if job.get('status')=='paused':
            paused.append({'job_id':jid,'processed':job.get('processed',0),'found':job.get('found',0),'log':job.get('log',[])[-3:]})
    return jsonify({'paused_jobs':paused})

@app.route('/upload',methods=['POST'])
def upload():
    if 'file' not in request.files: return jsonify({'error':'File nahi mila'}),400
    f=request.files['file']
    if f.filename=='': return jsonify({'error':'File select nahi ki'}),400
    filename=secure_filename(f.filename)
    filepath=os.path.join(app.config['UPLOAD_FOLDER'],filename)
    f.save(filepath)
    rows=[]; headers=[]
    try:
        if filename.endswith('.csv'):
            with open(filepath,encoding='utf-8-sig') as cf:
                reader=csv.DictReader(cf); rows=list(reader); headers=list(reader.fieldnames or [])
        else:
            import openpyxl; wb=openpyxl.load_workbook(filepath); ws=wb.active
            headers=[str(c.value or '').strip() for c in ws[1]]
            for row in ws.iter_rows(min_row=2,values_only=True): rows.append(dict(zip(headers,[str(v or '') for v in row])))
        with_website=sum(1 for r in rows if r.get('website','').strip())
        resume_options=[]
        for jid,job in jobs.items():
            if job.get('status')=='paused':
                resume_options.append({'job_id':jid,'last_row':job.get('processed',0),'found':job.get('found',0),'saved_at':'Last session'})
        return jsonify({'ok':True,'filename':filename,'filepath':filepath,'total':len(rows),'headers':headers,'with_website':with_website,'preview':rows[:5],'resume_options':resume_options})
    except Exception as e: return jsonify({'error':str(e)}),500

@app.route('/start',methods=['POST'])
def start_job():
    data=request.json
    filepath=data.get('filepath'); delay=float(data.get('delay',DELAY))
    resume_job_id=data.get('resume_job_id')
    if not filepath or not os.path.exists(filepath): return jsonify({'error':'File nahi mili'}),400
    rows=[]; fname=os.path.basename(filepath)
    try:
        if fname.endswith('.csv'):
            with open(filepath,encoding='utf-8-sig') as cf: rows=list(csv.DictReader(cf))
        else:
            import openpyxl; wb=openpyxl.load_workbook(filepath); ws=wb.active
            headers=[str(c.value or '') for c in ws[1]]
            for row in ws.iter_rows(min_row=2,values_only=True): rows.append(dict(zip(headers,[str(v or '') for v in row])))
    except Exception as e: return jsonify({'error':str(e)}),500
    start_from=0; existing_results=[]
    job_id=resume_job_id or str(uuid.uuid4())[:8]
    if resume_job_id and resume_job_id in jobs and jobs[resume_job_id].get('status')=='paused':
        paused=jobs[resume_job_id]
        start_from=paused.get('processed',0)
        existing_results=paused.get('results',[])
    elif resume_job_id:
        pcv=os.path.join('outputs',f'{resume_job_id}_progress.csv')
        lrf=os.path.join('outputs',f'{resume_job_id}_lastrow.txt')
        if os.path.exists(pcv) and os.path.exists(lrf):
            with open(lrf) as lf: start_from=int(lf.read().strip() or 0)
            with open(pcv,encoding='utf-8-sig') as cf: existing_results=list(csv.DictReader(cf))
    jobs[job_id]={'status':'queued','progress':0,'cancelled':False}
    t=threading.Thread(target=run_scraping_job,args=(job_id,rows,delay,start_from,existing_results))
    t.daemon=True; t.start()
    return jsonify({'job_id':job_id,'resumed_from':start_from})

@app.route('/status/<job_id>')
def job_status(job_id):
    if job_id not in jobs: return jsonify({'error':'Job nahi mila'}),404
    job=jobs[job_id]
    elapsed=int(time.time()-job.get('start_time',time.time()))
    processed=job.get('processed',0); total=job.get('total',1); found=job.get('found',0)
    eta_str=''
    if processed>0 and job['status']=='running':
        avg=elapsed/max(processed,1); remaining=int(avg*(total-processed))
        h,rem=divmod(remaining,3600); m,s=divmod(rem,60)
        eta_str=f"{h}h {m}m" if h>0 else f"{m}m {s}s"
    return jsonify({'status':job['status'],'progress':job.get('progress',0),'processed':processed,'total':total,'found':found,'eta':eta_str,'elapsed':elapsed,'log':job.get('log',[])[-20:],'has_csv':'out_csv' in job,'has_xlsx':'out_xlsx' in job,'preview':job.get('results',[])[-10:]})

@app.route('/cancel/<job_id>',methods=['POST'])
def cancel_job(job_id):
    if job_id in jobs: jobs[job_id]['cancelled']=True; return jsonify({'ok':True})
    return jsonify({'error':'Job nahi mila'}),404

@app.route('/download/<job_id>/<fmt>')
def download_file(job_id,fmt):
    if job_id not in jobs: return jsonify({'error':'Job nahi mila'}),404
    job=jobs[job_id]
    if fmt=='csv' and 'out_csv' in job:
        return send_file(job['out_csv'],as_attachment=True,download_name='emails_found.csv',mimetype='text/csv')
    elif fmt=='xlsx' and 'out_xlsx' in job:
        return send_file(job['out_xlsx'],as_attachment=True,download_name='emails_found.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return jsonify({'error':'File nahi mili'}),404

if __name__=='__main__':
    restore_saved_jobs()  # ← Check for saved jobs on startup
    print("\n"+"="*50)
    print("  EMAIL HUNTER PRO — Resume Edition")
    print("  http://localhost:5000")
    print("="*50+"\n")
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port, threaded=True, use_reloader=False)